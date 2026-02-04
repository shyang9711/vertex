#!/usr/bin/env python3
# kansai_payroll.py — tolerant to empty/missing cells
import sys, os, re
import tkinter as tk
from tkinter import filedialog, messagebox
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------- pick PDF ----------
def pick_pdf_path() -> str:
    if len(sys.argv) > 1 and os.path.isfile(sys.argv[1]):
        return os.path.abspath(sys.argv[1])
    root = tk.Tk(); root.withdraw()
    p = filedialog.askopenfilename(title="Select Payroll PDF", filetypes=[("PDF Files", "*.pdf")])
    if not p:
        try: messagebox.showerror("No file", "❌ No PDF file selected.")
        except Exception: pass
        sys.exit(1)
    return p

pdf_path = pick_pdf_path()
pdf_dir  = os.path.dirname(pdf_path)
base     = os.path.splitext(os.path.basename(pdf_path))[0]

# ---------- read & normalize ----------
doc = fitz.open(pdf_path)
lines = []
for page in doc:
    for ln in page.get_text().splitlines():
        s = (ln or "").strip()
        if s:
            s = re.sub(r"\s+", " ", s)  # normalize spaces
            lines.append(s)
doc.close()

# ---------- patterns ----------
SECTION_SET = {"SUSHI BAR", "KITCHEN", "SERVER"}
def norm_sec(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip()).upper()

money_re   = re.compile(r"\$([0-9][\d,]*(?:\.\d{1,2})?)")
rate_tok   = r"(?:\d{1,2}(?:\.\d{1,2})?)"                # 16.89 | 22
num_tok    = r"(?:\d+(?:\.\d{1,2})?)"                    # 44.4 | 0.30 | 88
big_num    = r"(?:\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?)"     # 8,333.33
detail_with_rate = re.compile(rf"^(?P<rate>{rate_tok})(?:\s+(?P<rest>.*))?$")

def is_header_or_footer(s: str) -> bool:
    return any(h in s for h in ("PAY PERIOD", "PAY DATE", "LineUp CPAs Mail", "TOTAL Memo", "https://"))

def is_name_line(s: str) -> bool:
    # Treat any digitless, non-$, non-section line with letters/commas as a name
    if any(ch.isdigit() for ch in s): return False
    if "$" in s: return False
    if norm_sec(s) in SECTION_SET: return False
    if is_header_or_footer(s): return False
    return re.search(r"[A-Za-z]", s) is not None

def to_float(x):
    if x is None or x == "": return 0.0
    return float(str(x).replace(",", ""))

def parse_sick(txt: str) -> float:
    m = re.search(r"(?:sick|병가)\s*(\d+(?:\.\d+)?)", txt, flags=re.I)
    return to_float(m.group(1)) if m else 0.0

# ---------- accumulators ----------
rows = {}                 # (name, section) -> totals
employee_salary = {}      # name -> salary (if standalone big number seen)
employee_memo   = {}      # name -> memo text
current_employee = None
current_section  = None
expecting_salary_or_section = False
debug = []  # optional diagnostics

def bucket_for(name: str, section: str):
    key = (name, section)
    if key not in rows:
        rows[key] = {
            "RATE": 0.0,
            "REG HRS": 0.0,
            "OVER TIME": 0.0,
            "DOUBLE TIME": 0.0,
            "TIP": 0.0,
            "CASH ADV": 0.0,
            "BONUS": 0.0,
            "SICK": 0.0,
        }
    return rows[key]

# ---------- parse ----------
for ln in lines:
    if is_header_or_footer(ln):
        continue

    sec_norm = norm_sec(ln)
    if sec_norm in SECTION_SET:
        current_section = sec_norm
        expecting_salary_or_section = False
        debug.append(("SECTION", current_section, ln))
        # ensure bucket exists for current employee+section even if details are empty
        if current_employee:
            bucket_for(current_employee, current_section)
        continue

    # name header (handle empty cells: name may be alone)
    if is_name_line(ln):
        current_employee = ln.strip()
        current_section  = None
        expecting_salary_or_section = True
        debug.append(("EMP_HDR", current_employee, ln))
        continue

    # standalone salary line right after name (no $)
    if expecting_salary_or_section and not money_re.search(ln) and re.fullmatch(big_num, ln):
        employee_salary[current_employee] = to_float(ln)
        debug.append(("SALARY", current_employee, ln))
        continue

    # Memo lines (e.g., "보고 x")
    if "보고" in ln and current_employee:
        employee_memo[current_employee] = ln
        debug.append(("MEMO", current_employee, ln))
        continue

    # need both to consume details; BUT tolerate missing details by keeping zeroes
    if not current_employee or not current_section:
        continue

    # detail line tolerant to empty cells:
    #   - may have only rate
    #   - may have only $tip
    #   - may have some hours, missing others
    mr = detail_with_rate.match(ln)
    if not mr:
        # fallback: if first token looks like rate, treat as detail
        toks = ln.split()
        if toks and re.fullmatch(rate_tok, toks[0]):
            rate_val = to_float(toks[0])
            rest = " ".join(toks[1:])
            class M: 
                def group(self, k): return str(rate_val) if k=="rate" else rest
            mr = M()
        else:
            # allow lines with no rate but with a $tip or bonus/cash-adv keywords
            if "$" not in ln and "보너스" not in ln and "빌려간돈" not in ln and "빼고" not in ln:
                continue
            # no rate → still attach amounts to bucket if present
            b = bucket_for(current_employee, current_section)
            # TIP-only lines like "$2118.24"
            mtip = money_re.findall(ln)
            if mtip:
                b["TIP"] += to_float(mtip[0])
            # explicit Korean cash-advance
            if ("빌려간돈" in ln) or ("빼고" in ln):
                m = re.search(rf"(?:빌려간돈\s*)?({big_num})\s*빼고", ln)
                if not m:
                    m = re.search(rf"빌려간돈\s*({big_num})", ln)
                if m:
                    b["CASH ADV"] += -to_float(m.group(1))
            # bonus
            if "보너스" in ln:
                m = re.search(rf"보너스\s*({big_num})", ln)
                if m:
                    b["BONUS"] += to_float(m.group(1))
            # sick
            b["SICK"] += parse_sick(ln)
            debug.append(("DETAIL_NO_RATE", f"{current_employee} / {current_section}", ln))
            continue

    # normal detail with a rate (even if rest is empty)
    rate = to_float(mr.group("rate"))
    rest = (mr.group("rest") or "").strip()

    b = bucket_for(current_employee, current_section)
    if b["RATE"] == 0.0:
        b["RATE"] = rate  # keep first seen rate; rows with only $tip keep RATE as prior or 0

    # token-by-token parse tolerant to empties:
    # - before first $: first three numeric tokens are REG / OT / DT (if present)
    # - any additional numeric before first $ becomes CASH ADV (subtract)
    # - at any time: first $... becomes TIP
    reg_set = ot_set = dt_set = False
    saw_dollar = False
    tokens = rest.split()
    for t in tokens:
        if t.startswith("$"):
            if not saw_dollar:
                b["TIP"] += to_float(t[1:])  # first dollar amount
                saw_dollar = True
            continue
        if re.fullmatch(num_tok, t):
            if not saw_dollar:
                if not reg_set:
                    b["REG HRS"] += to_float(t); reg_set = True; continue
                if not ot_set:
                    b["OVER TIME"] += to_float(t); ot_set = True; continue
                if not dt_set:
                    b["DOUBLE TIME"] += to_float(t); dt_set = True; continue
                # anything else before $ → treat as CASH ADV (subtract)
                b["CASH ADV"] += -to_float(t)
            else:
                # numbers after $ = printed totals; ignore
                pass
            continue
        # other tokens (words) are ignored here

    # explicit Korean phrases override/add
    if ("빌려간돈" in ln) or ("빼고" in ln):
        m = re.search(rf"(?:빌려간돈\s*)?({big_num})\s*빼고", ln)
        if not m:
            m = re.search(rf"빌려간돈\s*({big_num})", ln)
        if m:
            b["CASH ADV"] += -to_float(m.group(1))
    if "보너스" in ln:
        m = re.search(rf"보너스\s*({big_num})", ln)
        if m:
            b["BONUS"] += to_float(m.group(1))

    b["SICK"] += parse_sick(ln)
    debug.append(("DETAIL", f"{current_employee} / {current_section}", f"rate={rate} rest={rest}"))

# ---------- Excel ----------
wb = Workbook()
ws = wb.active
ws.title = "Payroll"

headers = [
    "SECTION","NAME","RATE","REG HRS","OVER TIME","DOUBLE TIME","TIP","CASH ADV","BONUS",
    "SALARY","SICK","HOURLY TOTAL","OVER TIME TOTAL","DOUBLE TIME TOTAL","SICK TOTAL","MEMO"
]
ws.append(headers)
for c in ws[1]: c.font = Font(bold=True)
ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions

col = {h: i+1 for i, h in enumerate(headers)}
GL = get_column_letter

# emit rows (ensure employees with salary but no sections still appear once)
emitted = False
for (name, section), agg in sorted(rows.items(), key=lambda kv: (kv[0][1], kv[0][0])):
    rate = agg["RATE"]; reg = agg["REG HRS"]; ot = agg["OVER TIME"]; dt = agg["DOUBLE TIME"]
    tip = agg["TIP"]; cadv = agg["CASH ADV"]; bon = agg["BONUS"]; sick = agg["SICK"]
    sal = employee_salary.get(name, 0.0)
    memo = employee_memo.get(name, "")

    ws.append([section, name, rate, reg, ot, dt, tip, cadv, bon, sal, sick, None, None, None, None, memo])
    r = ws.max_row
    cRATE = f"{GL(col['RATE'])}{r}"
    cREG  = f"{GL(col['REG HRS'])}{r}"
    cOT   = f"{GL(col['OVER TIME'])}{r}"
    cDT   = f"{GL(col['DOUBLE TIME'])}{r}"
    cSICK = f"{GL(col['SICK'])}{r}"

    ws.cell(row=r, column=col["HOURLY TOTAL"]).value       = f"={cRATE}*{cREG}"
    ws.cell(row=r, column=col["OVER TIME TOTAL"]).value    = f"={cRATE}*{cOT}*1.5"
    ws.cell(row=r, column=col["DOUBLE TIME TOTAL"]).value  = f"={cRATE}*{cDT}*2"
    ws.cell(row=r, column=col["SICK TOTAL"]).value         = f"={cRATE}*{cSICK}"
    emitted = True

# If no section details at all but we saw salaries/names, emit a salary-only row so you don't get 0 rows
if not emitted and employee_salary:
    for name, sal in employee_salary.items():
        ws.append(["", name, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, sal, 0.0, 0, 0, 0, 0, employee_memo.get(name, "")])

# highlight memo rows
hl = PatternFill(start_color="FFFBD17F", end_color="FFFBD17F", fill_type="solid")
for r in range(2, ws.max_row+1):
    if str(ws.cell(row=r, column=col["MEMO"]).value or "").strip():
        for c in range(1, len(headers)+1):
            ws.cell(row=r, column=c).fill = hl

# number formats
money_cols = ["RATE","TIP","CASH ADV","BONUS","SALARY","HOURLY TOTAL","OVER TIME TOTAL","DOUBLE TIME TOTAL","SICK TOTAL"]
hours_cols = ["REG HRS","OVER TIME","DOUBLE TIME","SICK"]
for name in money_cols:
    C = GL(col[name])
    for r in range(2, ws.max_row+1):
        ws[f"{C}{r}"].number_format = "#,##0.00"
for name in hours_cols:
    C = GL(col[name])
    for r in range(2, ws.max_row+1):
        ws[f"{C}{r}"].number_format = "0.00"

# auto-width
for c in range(1, len(headers)+1):
    L = GL(c); max_len = 0
    for r in range(1, ws.max_row+1):
        v = ws.cell(row=r, column=c).value
        if v is None: continue
        max_len = max(max_len, len(str(v)))
    ws.column_dimensions[L].width = min(max_len + 2, 60)

# optional DEBUG sheet (remove if you don't need it)
ws2 = wb.create_sheet("DEBUG")
ws2.append(["TYPE","WHO/SECTION","TEXT"])
for t,w,x in debug:
    ws2.append([t,w,x])

out_path = os.path.join(pdf_dir, base + "_parsed.xlsx")
wb.save(out_path)
try:
    messagebox.showinfo("Done", f"✅ Payroll parsed and saved to:\n{out_path}")
except Exception:
    pass
print(f"✅ Payroll parsed and saved to: {out_path}")
print(f"Parsed employees: {len({k[0] for k in rows.keys()})} | rows (name,section): {len(rows)}")
